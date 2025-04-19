import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def generate_report(db_params, filename):
    try:
        # Подключаемся к БД
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()

        # SQL-запрос
        query = """
        WITH object_total_costs AS (
            SELECT 
                o.id AS object_id,
                o.object_name,
                SUM(oe.object_estimates_price) AS object_total_cost
            FROM objects o
            JOIN object_estimates oe ON o.id = oe.object_id
            GROUP BY o.id, o.object_name
        ),
        ar_costs AS (
            SELECT 
                o.id AS object_id,
                SUM(le.local_estimates_price) AS total_ar_cost
            FROM objects o
            JOIN object_estimates oe ON o.id = oe.object_id
            JOIN local_estimates le ON oe.id = le.object_estimates_id
            WHERE le.name_local_estimate ILIKE ANY(ARRAY[
                '%Архитектурные решения%', 
                '%Кровля и кладка%',
                '%Витражи%',
                '%Отделочные работы%'
            ])
            GROUP BY o.id
        ),
        kr_costs AS (
            SELECT 
                o.id AS object_id,
                SUM(le.local_estimates_price) AS total_kr_cost
            FROM objects o
            JOIN object_estimates oe ON o.id = oe.object_id
            JOIN local_estimates le ON oe.id = le.object_estimates_id
            WHERE le.name_local_estimate ILIKE ANY(ARRAY[
                '%Конструктивные решения%',
                '%Конструкции железобетонные%',
                '%Конструктивные и объемно-планировочные решения%',
                '%Конструкции металлические%',
                '%Железобетонные,металлические конструкции%',
                '%Конструкции деревянные%',
                '%Железобетонные конструкции%'
            ])
            GROUP BY o.id
        )
        SELECT 
            otc.object_name,
            ROUND((COALESCE(ac.total_ar_cost, 0) / NULLIF(otc.object_total_cost, 0)) * 100, 2) AS ar_percentage,
            ROUND((COALESCE(kc.total_kr_cost, 0) / NULLIF(otc.object_total_cost, 0)) * 100, 2) AS kr_percentage
        FROM object_total_costs otc
        LEFT JOIN ar_costs ac ON otc.object_id = ac.object_id
        LEFT JOIN kr_costs kc ON otc.object_id = kc.object_id
        ORDER BY otc.object_id;
        """

        cursor.execute(query)
        results = cursor.fetchall()

        # Формируем DataFrame
        df = pd.DataFrame(results, columns=["Название объекта", "% АР", "% КР"])

        # Сохраняем в Excel
        df.to_excel(filename, index=False)

        # Настройка ширины столбцов по содержимому
        wb = load_workbook(filename)
        sheet = wb.active

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            # Проверяем длину каждого значения в столбце
            for cell in column:
                try:
                    value_length = len(str(cell.value))
                    if value_length > max_length:
                        max_length = value_length
                except:
                    pass

            # Устанавливаем ширину с небольшим запасом
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)

        print(f"Отчет успешно сохранен: {filename}")

    except Exception as e:
        print(f"Ошибка при генерации отчета: {str(e)}")
    finally:
        if conn:
            cursor.close()
            conn.close()